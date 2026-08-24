# =========================================================
# EXPORTACIONES A EXCEL
# Archivo: exports.py
# =========================================================

from io import BytesIO

import pandas as pd

import config
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------
# ESTILOS BASE DE EXCEL
# ---------------------------------------------------------
THIN_BORDER = Border(
    left=Side(style="thin", color="D9DEE5"),
    right=Side(style="thin", color="D9DEE5"),
    top=Side(style="thin", color="D9DEE5"),
    bottom=Side(style="thin", color="D9DEE5"),
)

TITLE_FILL = PatternFill(fill_type="solid", fgColor="1F2A44")
GLOBAL_TITLE_FILL = PatternFill(fill_type="solid", fgColor="000000")
HEADER_NEUTRAL_FILL = PatternFill(fill_type="solid", fgColor="1F2A44")
HEADER_ACTUAL_FILL = PatternFill(fill_type="solid", fgColor="0B5A7A")
HEADER_PLAN_FILL = PatternFill(fill_type="solid", fgColor="D4A017")

HEADER_PY_FILL = PatternFill(fill_type="solid", fgColor="0B5A7A")

LABEL_FILL = PatternFill(fill_type="solid", fgColor="F8FAFC")
TOTAL_FILL = PatternFill(fill_type="solid", fgColor="F3F6FA")
HIGHLIGHT_FILL = PatternFill(fill_type="solid", fgColor="E8F3E6")
HIGHLIGHT_LABEL_FILL = PatternFill(fill_type="solid", fgColor="DCEFD8")

TITLE_FONT = Font(color="FFFFFF", bold=True, size=12)
GLOBAL_TITLE_FONT = Font(color="FFFFFF", bold=True, size=14)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(color="1E1E1E", bold=False, size=10)
BODY_BOLD_FONT = Font(color="1E1E1E", bold=True, size=10)
NEGATIVE_FONT = Font(color="C0392B", bold=False, size=10)
NEGATIVE_BOLD_FONT = Font(color="C0392B", bold=True, size=10)

CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGNMENT = Alignment(horizontal="right", vertical="center")

# ---------------------------------------------------------
# CONFIGURACIÓN GENERAL
# ---------------------------------------------------------
DEFAULT_TABLE_SPACING_COLUMNS = 2
DEFAULT_BLOCK_SPACING_ROWS = 3
DEFAULT_SHEET_TITLE_ROW_HEIGHT = 22
DEFAULT_GLOBAL_TITLE_ROW_HEIGHT = 28
DEFAULT_HEADER_ROW_HEIGHT = 20
DEFAULT_BODY_ROW_HEIGHT = 19

PERCENT_COLUMNS = {"%Var VS Plan", "%Var VS Fcst", "%Var VS PY", "% GM", "Weight"}

NUMERIC_COLUMNS = {
    "Actual",
    "Plan",
    "Fcst",
    "PY",
    "Var VS Plan",
    "Var VS Fcst",
    "Var VS PY",
    "GSNR",
    "Gross Margin",
    "MTD Act",
    "MTD PY",
    "MTD Plan",
    "MTD Var vs PY",
    "MTD % Var vs PY",
    "MTD Var vs Plan",
    "MTD % Var vs Plan",
    "YTD Act",
    "YTD PY",
    "YTD Plan",
    "YTD Var vs PY",
    "YTD % Var vs PY",
    "YTD Var vs Plan",
    "YTD % Var vs Plan",
}

INTERNAL_COLUMNS_PREFIX = "__"


REPORT_METRIC_COLUMNS_EXECUTIVE_ORDER = [
    "Actual", "Plan", "Var VS Plan", "%Var VS Plan",
    "Fcst", "Var VS Fcst", "%Var VS Fcst",
    "PY", "Var VS PY", "%Var VS PY",
]


def reorder_export_metrics(df: pd.DataFrame) -> pd.DataFrame:
    """Mantiene dimensiones al inicio y acerca cada referencia a sus variaciones."""
    if df is None or df.empty:
        return df
    cols = list(df.columns)
    internal = [c for c in cols if is_internal_column(str(c))]
    metrics = [c for c in REPORT_METRIC_COLUMNS_EXECUTIVE_ORDER if c in cols]
    dimensions = [c for c in cols if c not in metrics and c not in internal]
    return df[dimensions + metrics + internal].copy()

# ---------------------------------------------------------
# CONFIGURACIÓN ESPECÍFICA REPORTE 4
# ---------------------------------------------------------
REPORT_4_VISIBLE_COLUMNS = [
    "TOP",
    "Client Name",
    "Cliente",
    "Actual",
    "Plan",
    "Var VS Plan",
    "%Var VS Plan",
    "Fcst",
    "Var VS Fcst",
    "%Var VS Fcst",
    "PY",
    "Var VS PY",
    "%Var VS PY",
]

REPORT_4_HIDDEN_EXPORT_COLUMNS = {"Grupo"}


# ---------------------------------------------------------
# HELPER DE PROGRESO
# ---------------------------------------------------------
def emit_progress(
    progress_callback,
    message: str,
    step: int,
    total_steps: int,
) -> None:
    """
    Envía una etapa real del proceso de exportación a la interfaz.

    El callback es opcional para conservar compatibilidad con las llamadas
    actuales. app.py será quien muestre estos mensajes mediante st.status.
    """
    if progress_callback is None:
        return

    progress_callback(
        message=message,
        step=step,
        total_steps=total_steps,
    )

# ---------------------------------------------------------
# HELPERS GENERALES
# ---------------------------------------------------------
def sanitize_sheet_name(sheet_name: str) -> str:
    invalid_chars = ["\\", "/", "*", "?", ":", "[", "]"]
    cleaned = str(sheet_name).strip()

    for char in invalid_chars:
        cleaned = cleaned.replace(char, "-")

    cleaned = cleaned[:31].strip()

    return cleaned or "Hoja"

def is_internal_column(column_name: str) -> bool:
    return str(column_name).startswith(INTERNAL_COLUMNS_PREFIX)

def sanitize_export_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    """
    Limpia cualquier DataFrame antes de exportarlo.

    Reglas generales:
    - Si no existe DataFrame, regresa uno vacío.
    - Quita columnas internas que empiezan con __.
    - Conserva el resto de columnas en el orden recibido.
    """
    if df is None:
        return pd.DataFrame()

    clean_df = reorder_export_metrics(df.copy())
    visible_columns = [
        col for col in clean_df.columns
        if not is_internal_column(str(col))
    ]

    return clean_df[visible_columns].copy()

def prepare_report_4_export_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    """
    Limpia las tablas del Reporte 4 para exportación.

    Reglas:
    - No exporta columnas internas.
    - Exporta TOP como primera columna y oculta únicamente Grupo.
    - Fuerza el mismo orden visible que la app: TOP, Client Name, Cliente y métricas.
    """
    if df is None:
        return pd.DataFrame(columns=REPORT_4_VISIBLE_COLUMNS)

    clean_df = df.copy()

    for column_name in REPORT_4_HIDDEN_EXPORT_COLUMNS:
        if column_name in clean_df.columns:
            clean_df = clean_df.drop(columns=[column_name])

    internal_columns = [col for col in clean_df.columns if is_internal_column(str(col))]
    clean_df = clean_df.drop(columns=internal_columns, errors="ignore")

    for column_name in REPORT_4_VISIBLE_COLUMNS:
        if column_name not in clean_df.columns:
            clean_df[column_name] = "" if column_name in {"Client Name", "Cliente"} else 0.0

    # Protección adicional: las llaves internas usadas para conservar clientes
    # sin código nunca deben aparecer en el Excel descargado.
    if "Cliente" in clean_df.columns:
        internal_client_mask = clean_df["Cliente"].astype(str).str.startswith(
            ("NAME_ONLY_", "PLAN_ONLY_")
        )
        clean_df.loc[internal_client_mask, "Cliente"] = "(blank)"

    return clean_df[REPORT_4_VISIBLE_COLUMNS].copy()

def get_first_column_name(df: pd.DataFrame) -> str | None:
    if df is None or df.empty or len(df.columns) == 0:
        return None
    return str(df.columns[0])

def safe_float(value):
    if value is None:
        return None

    try:
        numeric_value = float(value)
        if pd.isna(numeric_value):
            return None
        return numeric_value
    except (TypeError, ValueError):
        return None

def is_percent_column(column_name: str) -> bool:
    return str(column_name).strip() in PERCENT_COLUMNS

def is_numeric_column(column_name: str) -> bool:
    column_name = str(column_name).strip()
    return column_name in NUMERIC_COLUMNS or column_name in PERCENT_COLUMNS

def is_negative_number(value) -> bool:
    numeric_value = safe_float(value)
    return numeric_value is not None and numeric_value < 0

def estimate_column_width(series: pd.Series, header_text: str) -> int:
    max_length = len(str(header_text))

    for value in series.fillna("").astype(str).tolist():
        max_length = max(max_length, len(value))

    max_length = min(max_length + 2, 35)
    max_length = max(max_length, 12)

    return max_length

def excel_number_format_for_column(column_name: str) -> str:
    if is_percent_column(column_name):
        return "0.00%;[Red](0.00%);-"

    if is_numeric_column(column_name):
        return "#,##0;[Red](#,##0);-"

    return "@"

def scale_value_for_export(column_name: str, value):
    numeric_value = safe_float(value)

    if numeric_value is None:
        return None

    if is_percent_column(column_name):
        return numeric_value

    if is_numeric_column(column_name):
        return numeric_value / 1000

    return value

# ---------------------------------------------------------
# HELPERS DE ESTILO
# ---------------------------------------------------------
def apply_fill_if_present(cell, fill_style) -> None:
    if fill_style is not None:
        cell.fill = fill_style

def get_row_flags(original_df: pd.DataFrame, row_index: int) -> dict:
    if original_df is None or original_df.empty:
        return {
            "is_total": False,
            "is_grand_total": False,
            "is_highlight": False,
        }

    row = original_df.iloc[row_index]

    return {
        "is_total": bool(row.get("__is_total__", False)),
        "is_grand_total": bool(row.get("__is_grand_total__", False)),
        "is_highlight": bool(row.get("__is_highlight__", False)),
    }

def get_body_fill(column_name: str, row_flags: dict):
    is_total = row_flags.get("is_total", False)
    is_grand_total = row_flags.get("is_grand_total", False)
    is_highlight = row_flags.get("is_highlight", False)
    first_col = row_flags.get("first_col_name")

    if is_highlight:
        if column_name == first_col:
            return HIGHLIGHT_LABEL_FILL
        return HIGHLIGHT_FILL

    if is_total or is_grand_total:
        return TOTAL_FILL

    if column_name == first_col:
        return LABEL_FILL

    return None

def get_body_font(value, row_flags: dict, column_name: str | None = None):
    """
    Regla tipográfica:
    - Totales, subtotales y resaltados: toda la fila en negritas.
    - En Reporte 4, Client Name y Cliente: en negritas.
    - Resultados normales: sin negritas.
    - Negativos normales: rojo sin negritas.
    - Negativos en totales/subtotales: rojo y negritas.
    """
    is_total = row_flags.get("is_total", False)
    is_grand_total = row_flags.get("is_grand_total", False)
    is_highlight = row_flags.get("is_highlight", False)
    is_special_row = is_total or is_grand_total or is_highlight
    clean_column_name = str(column_name or "").strip()

    if is_negative_number(value):
        return NEGATIVE_BOLD_FONT if is_special_row else NEGATIVE_FONT

    if is_special_row:
        return BODY_BOLD_FONT

    if clean_column_name in {"Client Name", "Cliente"}:
        return BODY_BOLD_FONT

    return BODY_FONT

def get_header_fill(column_name: str):
    column_name = str(column_name).strip()

    if column_name == "Actual":
        return HEADER_ACTUAL_FILL
    if column_name == "Plan":
        return HEADER_PLAN_FILL
    if column_name == "PY":
        return HEADER_PY_FILL

    return HEADER_NEUTRAL_FILL

def write_global_title(
    worksheet,
    report_title: str | None,
    start_row: int = 1,
    start_col: int = 1,
    width: int = 8,
) -> int:
    """
    Escribe un encabezado general del reporte en Excel.
    Devuelve la siguiente fila disponible.
    """
    if not report_title:
        return start_row

    worksheet.merge_cells(
        start_row=start_row,
        start_column=start_col,
        end_row=start_row,
        end_column=start_col + width - 1,
    )

    title_cell = worksheet.cell(row=start_row, column=start_col, value=report_title)
    title_cell.font = GLOBAL_TITLE_FONT
    title_cell.alignment = CENTER_ALIGNMENT
    title_cell.border = THIN_BORDER
    title_cell.fill = GLOBAL_TITLE_FILL

    worksheet.row_dimensions[start_row].height = DEFAULT_GLOBAL_TITLE_ROW_HEIGHT

    return start_row + 2

# ---------------------------------------------------------
# ESCRITURA DE TABLAS
# ---------------------------------------------------------
def write_table_to_worksheet(
    worksheet,
    original_df: pd.DataFrame,
    start_row: int,
    start_col: int,
    table_title: str,
) -> tuple[int, int]:
    export_df = sanitize_export_dataframe(original_df)

    if export_df is None or export_df.empty:
        title_cell = worksheet.cell(row=start_row, column=start_col, value=table_title)
        title_cell.font = TITLE_FONT
        title_cell.alignment = LEFT_ALIGNMENT
        title_cell.border = THIN_BORDER
        apply_fill_if_present(title_cell, TITLE_FILL)

        info_cell = worksheet.cell(
            row=start_row + 1,
            column=start_col,
            value="Sin información disponible",
        )
        info_cell.font = BODY_FONT
        info_cell.alignment = LEFT_ALIGNMENT
        info_cell.border = THIN_BORDER

        return start_row + 1, start_col

    num_cols = len(export_df.columns)
    first_col_name = get_first_column_name(export_df)

    title_row = start_row
    header_row = start_row + 1
    body_start_row = start_row + 2

    worksheet.merge_cells(
        start_row=title_row,
        start_column=start_col,
        end_row=title_row,
        end_column=start_col + num_cols - 1,
    )

    title_cell = worksheet.cell(row=title_row, column=start_col, value=table_title)
    title_cell.font = TITLE_FONT
    title_cell.alignment = LEFT_ALIGNMENT
    title_cell.border = THIN_BORDER
    apply_fill_if_present(title_cell, TITLE_FILL)

    worksheet.row_dimensions[title_row].height = DEFAULT_SHEET_TITLE_ROW_HEIGHT
    worksheet.row_dimensions[header_row].height = DEFAULT_HEADER_ROW_HEIGHT

    for offset, column_name in enumerate(export_df.columns):
        current_col = start_col + offset
        header_cell = worksheet.cell(
            row=header_row,
            column=current_col,
            value=str(column_name),
        )
        header_cell.font = HEADER_FONT
        header_cell.alignment = LEFT_ALIGNMENT if offset == 0 else CENTER_ALIGNMENT
        header_cell.border = THIN_BORDER
        apply_fill_if_present(header_cell, get_header_fill(column_name))

    for row_idx in range(len(export_df)):
        excel_row = body_start_row + row_idx
        worksheet.row_dimensions[excel_row].height = DEFAULT_BODY_ROW_HEIGHT

        row_flags = get_row_flags(original_df, row_idx)
        row_flags["first_col_name"] = first_col_name

        for col_idx, column_name in enumerate(export_df.columns):
            excel_col = start_col + col_idx
            raw_value = export_df.iloc[row_idx, col_idx]

            output_value = raw_value
            if is_numeric_column(column_name):
                output_value = scale_value_for_export(column_name, raw_value)
                if output_value is None:
                    output_value = "-"

            cell = worksheet.cell(
                row=excel_row,
                column=excel_col,
                value=output_value,
            )

            cell.border = THIN_BORDER
            cell.font = get_body_font(raw_value, row_flags, str(column_name))

            fill_style = get_body_fill(str(column_name), row_flags)
            apply_fill_if_present(cell, fill_style)

            if str(column_name).strip() in {"Client Name", "Cliente"} or col_idx == 0:
                cell.alignment = LEFT_ALIGNMENT
            else:
                cell.alignment = RIGHT_ALIGNMENT if is_numeric_column(column_name) else CENTER_ALIGNMENT

            if is_numeric_column(column_name):
                cell.number_format = excel_number_format_for_column(column_name)

    for offset, column_name in enumerate(export_df.columns):
        current_col = start_col + offset
        column_letter = get_column_letter(current_col)
        width = estimate_column_width(export_df[column_name], str(column_name))
        worksheet.column_dimensions[column_letter].width = width

    last_row = body_start_row + len(export_df) - 1
    last_col = start_col + num_cols - 1

    return last_row, last_col

def write_two_tables_side_by_side(
    worksheet,
    left_df: pd.DataFrame,
    right_df: pd.DataFrame,
    left_title: str,
    right_title: str,
    start_row: int,
    start_col: int = 1,
    spacing_columns: int = DEFAULT_TABLE_SPACING_COLUMNS,
) -> tuple[int, int]:
    left_visible_df = sanitize_export_dataframe(left_df)
    left_cols = max(len(left_visible_df.columns), 1)

    left_last_row, left_last_col = write_table_to_worksheet(
        worksheet=worksheet,
        original_df=left_df,
        start_row=start_row,
        start_col=start_col,
        table_title=left_title,
    )

    right_start_col = start_col + left_cols + spacing_columns

    right_last_row, right_last_col = write_table_to_worksheet(
        worksheet=worksheet,
        original_df=right_df,
        start_row=start_row,
        start_col=right_start_col,
        table_title=right_title,
    )

    return max(left_last_row, right_last_row), max(left_last_col, right_last_col)

# ---------------------------------------------------------
# HELPERS DE EXCEL
# ---------------------------------------------------------
def create_excel_writer_buffer():
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine="openpyxl")
    return output, writer

def build_excel_bytes_from_writer(output: BytesIO, writer: pd.ExcelWriter) -> bytes:
    writer.close()
    output.seek(0)
    return output.getvalue()

def ensure_sheet_exists(writer, sheet_name: str):
    safe_sheet_name = sanitize_sheet_name(sheet_name)
    if safe_sheet_name not in writer.book.sheetnames:
        writer.book.create_sheet(title=safe_sheet_name)
    return writer.book[safe_sheet_name]

def remove_default_sheet_if_needed(writer) -> None:
    if "Sheet" in writer.book.sheetnames and len(writer.book.sheetnames) > 1:
        default_sheet = writer.book["Sheet"]
        writer.book.remove(default_sheet)

def get_report_title_from_tables(tables: dict | None, fallback: str | None = None) -> str | None:
    if isinstance(tables, dict):
        return tables.get("report_title") or tables.get("__report_title__") or fallback
    return fallback


# ---------------------------------------------------------
# EXPORTACIÓN BASE MTD / BTS
# ---------------------------------------------------------
def build_base_mtd_excel_bytes(
    client_table_df: pd.DataFrame,
    sku_table_df: pd.DataFrame,
    bts_table_df: pd.DataFrame,
    plan_summary_df: pd.DataFrame | None = None,
    report_title: str | None = None,
    sheet_name: str | None = None,
    progress_callback=None,
) -> bytes:
    """
    Construye el Excel individual de Base MTD.

    Incluye únicamente:
    - Comparativo MTD/YTD contra Plan Cliente.
    - Comparativo MTD/YTD contra Plan SKU.
    - Tabla BTS.

    Nota:
    plan_summary_df se conserva como parámetro para no romper las llamadas
    existentes desde app.py, pero ya no se escribe en el Excel.
    """
    total_steps = 5
    emit_progress(progress_callback, "Preparando el archivo de Base MTD", 1, total_steps)

    output, writer = create_excel_writer_buffer()

    emit_progress(progress_callback, "Creando la hoja de Base MTD", 2, total_steps)
    worksheet = ensure_sheet_exists(
        writer,
        sheet_name or getattr(config, "EXPORT_SHEET_BASE_MTD", "Base MTD"),
    )

    current_row = write_global_title(
        worksheet=worksheet,
        report_title=report_title,
        start_row=1,
        start_col=1,
        width=8,
    )

    emit_progress(progress_callback, "Escribiendo comparativos de Plan Cliente y Plan SKU", 3, total_steps)
    block_last_row, _ = write_two_tables_side_by_side(
        worksheet=worksheet,
        left_df=client_table_df,
        right_df=sku_table_df,
        left_title=getattr(config, "BASE_MTD_CLIENT_TABLE_TITLE", "Base MTD vs Plan Cliente"),
        right_title=getattr(config, "BASE_MTD_SKU_TABLE_TITLE", "Base MTD vs Plan SKU"),
        start_row=current_row,
    )

    current_row = block_last_row + DEFAULT_BLOCK_SPACING_ROWS

    emit_progress(progress_callback, "Escribiendo la tabla BTS", 4, total_steps)
    write_table_to_worksheet(
        worksheet=worksheet,
        original_df=bts_table_df,
        start_row=current_row,
        start_col=1,
        table_title=getattr(config, "BASE_MTD_BTS_TABLE_TITLE", "Back To School (BTS)"),
    )

    emit_progress(progress_callback, "Finalizando el archivo Excel de Base MTD", 5, total_steps)
    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)


def write_base_mtd_tables_to_workbook(
    worksheet,
    base_mtd_tables: dict,
) -> None:
    """
    Escribe Base MTD dentro del archivo global de reportes.
    """
    current_row = write_global_title(
        worksheet=worksheet,
        report_title=get_report_title_from_tables(base_mtd_tables),
        start_row=1,
        start_col=1,
        width=8,
    )

    block_last_row, _ = write_two_tables_side_by_side(
        worksheet=worksheet,
        left_df=base_mtd_tables.get("client_table"),
        right_df=base_mtd_tables.get("sku_table"),
        left_title=getattr(config, "BASE_MTD_CLIENT_TABLE_TITLE", "Base MTD vs Plan Cliente"),
        right_title=getattr(config, "BASE_MTD_SKU_TABLE_TITLE", "Base MTD vs Plan SKU"),
        start_row=current_row,
    )

    current_row = block_last_row + DEFAULT_BLOCK_SPACING_ROWS

    write_table_to_worksheet(
        worksheet=worksheet,
        original_df=base_mtd_tables.get("bts_table"),
        start_row=current_row,
        start_col=1,
        table_title=getattr(config, "BASE_MTD_BTS_TABLE_TITLE", "Back To School (BTS)"),
    )

# ---------------------------------------------------------
# EXPORTACIÓN REPORTE 1
# ---------------------------------------------------------
def build_report_1_excel_bytes(
    mtd_without_kens_df: pd.DataFrame,
    ytd_without_kens_df: pd.DataFrame,
    report_title: str | None = None,
    sheet_name: str = "Reporte 1",
    progress_callback=None,
) -> bytes:
    total_steps = 4
    emit_progress(progress_callback, "Preparando la exportación de Oficina de ventas", 1, total_steps)

    output, writer = create_excel_writer_buffer()

    emit_progress(progress_callback, "Creando la hoja del Reporte 1", 2, total_steps)
    worksheet = ensure_sheet_exists(writer, sheet_name)

    current_row = write_global_title(
        worksheet=worksheet,
        report_title=report_title,
        start_row=1,
        start_col=1,
        width=8,
    )

    emit_progress(progress_callback, "Escribiendo las tablas MTD y YTD de Oficina de ventas", 3, total_steps)
    write_two_tables_side_by_side(
        worksheet=worksheet,
        left_df=mtd_without_kens_df,
        right_df=ytd_without_kens_df,
        left_title="MTD Oficina de ventas",
        right_title="YTD Oficina de ventas",
        start_row=current_row,
    )

    emit_progress(progress_callback, "Finalizando el archivo Excel del Reporte 1", 4, total_steps)
    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)

# ---------------------------------------------------------
# EXPORTACIÓN REPORTE 2 - SEGMENT X REGION
# ---------------------------------------------------------
def build_report_2_segment_excel_bytes(
    mtd_segment_df: pd.DataFrame,
    ytd_segment_df: pd.DataFrame,
    report_title: str | None = None,
    sheet_name: str = "Reporte 2 - Segment",
    progress_callback=None,
) -> bytes:
    total_steps = 4
    emit_progress(progress_callback, "Preparando la exportación de Segment x Region", 1, total_steps)

    output, writer = create_excel_writer_buffer()

    emit_progress(progress_callback, "Creando la hoja de Segment x Region", 2, total_steps)
    worksheet = ensure_sheet_exists(writer, sheet_name)

    current_row = write_global_title(
        worksheet=worksheet,
        report_title=report_title,
        start_row=1,
        start_col=1,
        width=8,
    )

    emit_progress(progress_callback, "Escribiendo las tablas MTD y YTD de Segment x Region", 3, total_steps)
    write_two_tables_side_by_side(
        worksheet=worksheet,
        left_df=mtd_segment_df,
        right_df=ytd_segment_df,
        left_title="MTD Segment x Region",
        right_title="YTD Segment x Region",
        start_row=current_row,
    )

    emit_progress(progress_callback, "Finalizando el archivo Excel de Segment x Region", 4, total_steps)
    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)

# ---------------------------------------------------------
# EXPORTACIÓN REPORTE 2 - CATEGORY
# ---------------------------------------------------------
def build_report_2_category_excel_bytes(
    mtd_category_df: pd.DataFrame,
    ytd_category_df: pd.DataFrame,
    report_title: str | None = None,
    sheet_name: str = "Reporte 2 - Category",
    progress_callback=None,
) -> bytes:
    total_steps = 4
    emit_progress(progress_callback, "Preparando la exportación de Category", 1, total_steps)

    output, writer = create_excel_writer_buffer()

    emit_progress(progress_callback, "Creando la hoja de Category", 2, total_steps)
    worksheet = ensure_sheet_exists(writer, sheet_name)

    current_row = write_global_title(
        worksheet=worksheet,
        report_title=report_title,
        start_row=1,
        start_col=1,
        width=9,
    )

    emit_progress(progress_callback, "Escribiendo las tablas MTD y YTD de Category", 3, total_steps)
    write_two_tables_side_by_side(
        worksheet=worksheet,
        left_df=mtd_category_df,
        right_df=ytd_category_df,
        left_title="MTD Category",
        right_title="YTD Category",
        start_row=current_row,
    )

    emit_progress(progress_callback, "Finalizando el archivo Excel de Category", 4, total_steps)
    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)

# ---------------------------------------------------------
# EXPORTACIÓN REPORTE 3
# ---------------------------------------------------------
def build_report_3_excel_bytes(
    mtd_channel_df: pd.DataFrame,
    ytd_channel_df: pd.DataFrame,
    report_title: str | None = None,
    sheet_name: str = "Reporte 3",
    progress_callback=None,
) -> bytes:
    total_steps = 4
    emit_progress(progress_callback, "Preparando la exportación de Channel", 1, total_steps)

    output, writer = create_excel_writer_buffer()

    emit_progress(progress_callback, "Creando la hoja del Reporte 3", 2, total_steps)
    worksheet = ensure_sheet_exists(writer, sheet_name)

    current_row = write_global_title(
        worksheet=worksheet,
        report_title=report_title,
        start_row=1,
        start_col=1,
        width=8,
    )

    emit_progress(progress_callback, "Escribiendo las tablas MTD y YTD de Channel", 3, total_steps)
    write_two_tables_side_by_side(
        worksheet=worksheet,
        left_df=mtd_channel_df,
        right_df=ytd_channel_df,
        left_title="MTD Channel",
        right_title="YTD Channel",
        start_row=current_row,
    )

    emit_progress(progress_callback, "Finalizando el archivo Excel del Reporte 3", 4, total_steps)
    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)

# ---------------------------------------------------------
# EXPORTACIÓN REPORTE 4
# ---------------------------------------------------------
def build_report_4_excel_bytes(
    mtd_top_clients_df: pd.DataFrame,
    ytd_top_clients_df: pd.DataFrame,
    report_title: str | None = None,
    sheet_name: str = "Reporte 4",
    progress_callback=None,
) -> bytes:
    total_steps = 4
    emit_progress(progress_callback, "Preparando la exportación del Ranking de Clientes", 1, total_steps)

    output, writer = create_excel_writer_buffer()

    emit_progress(progress_callback, "Creando la hoja del Reporte 4", 2, total_steps)
    worksheet = ensure_sheet_exists(writer, sheet_name)

    mtd_export_df = prepare_report_4_export_dataframe(mtd_top_clients_df)
    ytd_export_df = prepare_report_4_export_dataframe(ytd_top_clients_df)

    current_row = write_global_title(
        worksheet=worksheet,
        report_title=report_title,
        start_row=1,
        start_col=1,
        width=len(REPORT_4_VISIBLE_COLUMNS),
    )

    emit_progress(progress_callback, "Escribiendo las tablas MTD y YTD del Ranking", 3, total_steps)
    write_two_tables_side_by_side(
        worksheet=worksheet,
        left_df=mtd_export_df,
        right_df=ytd_export_df,
        left_title="MTD Ranking Clients",
        right_title="YTD Ranking Clients",
        start_row=current_row,
    )

    emit_progress(progress_callback, "Finalizando el archivo Excel del Reporte 4", 4, total_steps)
    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)

# ---------------------------------------------------------
# EXPORTACIÓN DASHBOARD EJECUTIVO
# ---------------------------------------------------------
# La hoja Dashboard NO reutiliza el diseño azul/dorado de los reportes
# individuales. Se construye celda por celda para conservar la apariencia
# limpia del Dashboard mostrado en Streamlit.

DASHBOARD_RED = "E60023"
DASHBOARD_BLACK = "111111"
DASHBOARD_TEXT = "000000"
DASHBOARD_NEGATIVE = "C0392B"
DASHBOARD_TITLE_GRAY = "D9DDE3"
DASHBOARD_GSNR_GRAY = "D9DDE3"
DASHBOARD_BTS_GRAY = "F1F3F5"
DASHBOARD_TOTAL_GRAY = "F1F3F5"
DASHBOARD_GRAND_TOTAL_GREEN = "E8F3E6"
DASHBOARD_WHITE = "FFFFFF"

DASHBOARD_NO_BORDER = Border()
DASHBOARD_HEADER_BORDER = Border(
    bottom=Side(style="double", color=DASHBOARD_BLACK),
)
DASHBOARD_KPI_HEADER_BORDER = Border(
    bottom=Side(style="medium", color=DASHBOARD_BLACK),
)
DASHBOARD_TOTAL_BORDER = Border(
    top=Side(style="double", color=DASHBOARD_BLACK),
)

DASHBOARD_TITLE_FONT = Font(
    name="Calibri",
    size=18,
    bold=True,
    color=DASHBOARD_RED,
)
DASHBOARD_SECTION_FONT = Font(
    name="Calibri",
    size=11,
    bold=True,
    color=DASHBOARD_TEXT,
)
DASHBOARD_HEADER_FONT = Font(
    name="Calibri",
    size=9,
    bold=True,
    color=DASHBOARD_TEXT,
)
DASHBOARD_BODY_FONT = Font(
    name="Calibri",
    size=9,
    bold=False,
    color=DASHBOARD_TEXT,
)
DASHBOARD_BODY_BOLD_FONT = Font(
    name="Calibri",
    size=9,
    bold=True,
    color=DASHBOARD_TEXT,
)
DASHBOARD_NEGATIVE_FONT = Font(
    name="Calibri",
    size=9,
    bold=True,
    color=DASHBOARD_NEGATIVE,
)
DASHBOARD_PERIOD_LABEL_FONT = Font(
    name="Calibri",
    size=10,
    bold=True,
    color=DASHBOARD_RED,
)
DASHBOARD_PERIOD_VALUE_FONT = Font(
    name="Calibri",
    size=10,
    bold=True,
    color=DASHBOARD_TEXT,
)

DASHBOARD_NUMBER_FORMAT = '#,##0;[Red](#,##0);-'
DASHBOARD_PERCENT_FORMAT = '0.00%;[Red](0.00%);-'


def _dashboard_find_period_row(df: pd.DataFrame | None, period: str):
    if df is None or df.empty or "Periodo" not in df.columns:
        return None

    matches = df[
        df["Periodo"].astype(str).str.strip().str.upper()
        == str(period).strip().upper()
    ]
    if matches.empty:
        return None

    return matches.iloc[0]


def _dashboard_kpi_dataframe(
    client_table: pd.DataFrame | None,
    bts_table: pd.DataFrame | None,
    period: str,
    bts_label: str,
) -> pd.DataFrame:
    """
    Construye KPI Dashboard con Actual, Plan, Forecast y PY.
    """
    gsnr_row = _dashboard_find_period_row(client_table, period)
    bts_row = _dashboard_find_period_row(bts_table, period)

    def get_value(row, column, default=None):
        if row is None or column not in row.index:
            return default
        return row.get(column, default)

    actual = get_value(gsnr_row, "Actual")
    plan = get_value(gsnr_row, "Plan")
    fcst = get_value(gsnr_row, "Fcst")
    py = get_value(gsnr_row, "PY")

    actual_numeric = safe_float(actual)
    plan_numeric = safe_float(plan)
    fcst_numeric = safe_float(fcst)
    py_numeric = safe_float(py)

    achievement_plan = (
        None
        if actual_numeric is None or plan_numeric in (None, 0)
        else actual_numeric / plan_numeric
    )
    achievement_fcst = (
        None
        if actual_numeric is None or fcst_numeric in (None, 0)
        else actual_numeric / fcst_numeric
    )
    achievement_py = (
        None
        if actual_numeric is None or py_numeric in (None, 0)
        else actual_numeric / py_numeric
    )

    return pd.DataFrame([
        {
            "KPI": "GSNR",
            "Actual": actual,
            "Plan": plan,
            "Fcst": fcst,
            "PY": py,
            "Var VS Plan": get_value(gsnr_row, "Var VS Plan"),
            "%Var VS Plan": get_value(gsnr_row, "%Var VS Plan"),
            "Var VS Fcst": get_value(gsnr_row, "Var VS Fcst"),
            "%Var VS Fcst": get_value(gsnr_row, "%Var VS Fcst"),
            "Var VS PY": get_value(gsnr_row, "Var VS PY"),
            "%Var VS PY": get_value(gsnr_row, "%Var VS PY"),
        },
        {
            "KPI": "% achievement",
            "Actual": None,
            "Plan": None,
            "Fcst": None,
            "PY": None,
            "Var VS Plan": achievement_plan,
            "%Var VS Plan": None,
            "Var VS Fcst": achievement_fcst,
            "%Var VS Fcst": None,
            "Var VS PY": achievement_py,
            "%Var VS PY": None,
        },
        {
            "KPI": bts_label,
            "Actual": get_value(bts_row, "Actual"),
            "Plan": None,
            "Fcst": None,
            "PY": get_value(bts_row, "PY"),
            "Var VS Plan": None,
            "%Var VS Plan": 0.0,
            "Var VS Fcst": None,
            "%Var VS Fcst": None,
            "Var VS PY": get_value(bts_row, "Var VS PY"),
            "%Var VS PY": get_value(bts_row, "%Var VS PY"),
        },
    ])


def _dashboard_is_percent_column(column_name: str) -> bool:
    clean_name = str(column_name).strip()
    return clean_name.startswith("%") or clean_name in {
        "%Var VS Plan",
        "%Var VS PY",
        "% GM",
        "Weight",
    }


def _dashboard_is_numeric_column(column_name: str) -> bool:
    clean_name = str(column_name).strip()
    return clean_name in NUMERIC_COLUMNS or _dashboard_is_percent_column(clean_name)


def _dashboard_scale_value(column_name: str, value):
    numeric_value = safe_float(value)
    if numeric_value is None:
        return None

    if _dashboard_is_percent_column(column_name):
        return numeric_value

    if _dashboard_is_numeric_column(column_name):
        return numeric_value / 1000

    return value


def _dashboard_set_numeric_style(cell, column_name: str, raw_value) -> None:
    if _dashboard_is_percent_column(column_name):
        cell.number_format = DASHBOARD_PERCENT_FORMAT
    elif _dashboard_is_numeric_column(column_name):
        cell.number_format = DASHBOARD_NUMBER_FORMAT

    if is_negative_number(raw_value):
        cell.font = DASHBOARD_NEGATIVE_FONT



def _dashboard_recalculate_metrics(
    row: dict,
    actual,
    plan,
    py,
    fcst=None,
) -> dict:
    """
    Recalcula las métricas consolidadas del Dashboard incluyendo Forecast.
    """
    result = dict(row or {})

    actual_num = safe_float(actual) or 0.0
    plan_num = safe_float(plan) or 0.0
    fcst_num = safe_float(fcst) or 0.0
    py_num = safe_float(py) or 0.0

    result["Actual"] = actual_num
    result["Plan"] = plan_num
    result["Fcst"] = fcst_num
    result["PY"] = py_num

    result["Var VS Plan"] = actual_num - plan_num
    result["%Var VS Plan"] = (
        (actual_num - plan_num) / plan_num
        if plan_num != 0
        else None
    )

    result["Var VS Fcst"] = actual_num - fcst_num
    result["%Var VS Fcst"] = (
        (actual_num - fcst_num) / fcst_num
        if fcst_num != 0
        else None
    )

    result["Var VS PY"] = actual_num - py_num
    result["%Var VS PY"] = (
        (actual_num - py_num) / py_num
        if py_num != 0
        else None
    )

    return result


def _dashboard_category_order(
    mtd_df: pd.DataFrame | None,
    ytd_df: pd.DataFrame | None,
) -> list[str]:
    """Obtiene la unión ordenada de categorías usada por Monthly y YTD."""
    ordered: list[str] = []
    for df in (mtd_df, ytd_df):
        if df is None or df.empty or "Category" not in df.columns:
            continue
        for _, row in df.iterrows():
            if bool(row.get("__is_grand_total__", False)):
                continue
            label = str(row.get("Category", "")).strip()
            if not label or label.lower() in {"total mexico", "total general", "grand total"}:
                continue
            if label not in ordered:
                ordered.append(label)
    return ordered


def _dashboard_aggregate_category_table(
    df: pd.DataFrame | None,
    ordered_categories: list[str],
) -> pd.DataFrame:
    """
    Convierte el Reporte de Category detallado por material en la vista
    ejecutiva del Dashboard: una sola fila por Category.
    """
    visible_columns = [
        "Category", "Actual", "Plan", "Fcst", "PY", "Var VS Plan",
        "%Var VS Plan", "Var VS Fcst", "%Var VS Fcst",
        "Var VS PY", "%Var VS PY",
        "__is_total__", "__is_grand_total__", "__is_highlight__",
    ]

    if df is None or df.empty:
        rows = []
        for category in ordered_categories:
            rows.append({
                "Category": category,
                "Actual": 0.0,
                "Plan": 0.0,
                "PY": 0.0,
                "Var VS Plan": 0.0,
                "%Var VS Plan": None,
                "Var VS PY": 0.0,
                "%Var VS PY": None,
                "__is_total__": False,
                "__is_grand_total__": False,
                "__is_highlight__": False,
            })
        return pd.DataFrame(rows, columns=visible_columns)

    totals_by_category: dict[str, dict] = {}
    detail_by_category: dict[str, list[dict]] = {}
    grand_total: dict | None = None

    for _, source_row in df.iterrows():
        row = dict(source_row)
        category = str(row.get("Category", "")).strip()
        normalized = category.lower()

        if bool(row.get("__is_grand_total__", False)) or normalized in {
            "total mexico", "total general", "grand total"
        }:
            grand_total = row
            continue
        if not category:
            continue
        if bool(row.get("__is_total__", False)):
            totals_by_category[category] = row
        else:
            detail_by_category.setdefault(category, []).append(row)

    # Si el reporte no trae una fila total por categoría, suma el detalle.
    for category, detail_rows in detail_by_category.items():
        if category in totals_by_category:
            continue
        actual = sum((safe_float(r.get("Actual")) or 0.0) for r in detail_rows)
        plan = sum((safe_float(r.get("Plan")) or 0.0) for r in detail_rows)
        fcst = sum((safe_float(r.get("Fcst")) or 0.0) for r in detail_rows)
        py = sum((safe_float(r.get("PY")) or 0.0) for r in detail_rows)
        totals_by_category[category] = _dashboard_recalculate_metrics(
            detail_rows[0] if detail_rows else {},
            actual,
            plan,
            py,
            fcst=fcst,
        )

    rows: list[dict] = []
    for category in ordered_categories:
        row = dict(totals_by_category.get(category, {}))
        if not row:
            row = _dashboard_recalculate_metrics({}, 0.0, 0.0, 0.0)
        row["Category"] = category
        row["__is_total__"] = False
        row["__is_grand_total__"] = False
        row["__is_highlight__"] = False
        rows.append(row)

    if grand_total is None:
        actual = sum((safe_float(r.get("Actual")) or 0.0) for r in rows)
        plan = sum((safe_float(r.get("Plan")) or 0.0) for r in rows)
        fcst = sum((safe_float(r.get("Fcst")) or 0.0) for r in rows)
        py = sum((safe_float(r.get("PY")) or 0.0) for r in rows)
        grand_total = _dashboard_recalculate_metrics(
            {},
            actual,
            plan,
            py,
            fcst=fcst,
        )

    grand_total = dict(grand_total)
    grand_total["Category"] = "Total Mexico"
    grand_total["__is_total__"] = True
    grand_total["__is_grand_total__"] = False
    grand_total["__is_highlight__"] = False
    rows.append(grand_total)

    result = pd.DataFrame(rows)
    for col in visible_columns:
        if col not in result.columns:
            result[col] = False if col.startswith("__") else None
    return result[visible_columns].copy()


def _dashboard_prepare_category_pair(
    mtd_df: pd.DataFrame | None,
    ytd_df: pd.DataFrame | None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    order = _dashboard_category_order(mtd_df, ytd_df)
    return (
        _dashboard_aggregate_category_table(mtd_df, order),
        _dashboard_aggregate_category_table(ytd_df, order),
    )


DASHBOARD_RANKING_VISIBLE_COLUMNS = [
    "Client Name",
    "Actual",
    "Plan",
    "Fcst",
    "PY",
    "Var VS Plan",
    "%Var VS Plan",
    "Var VS Fcst",
    "%Var VS Fcst",
    "Var VS PY",
    "%Var VS PY",
]


def _dashboard_prepare_ranking_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    """Prepara el Ranking exclusivo del Dashboard, sin la columna Cliente."""
    if df is None:
        return pd.DataFrame(columns=DASHBOARD_RANKING_VISIBLE_COLUMNS)

    clean = df.copy()
    clean = clean.drop(
        columns=list(REPORT_4_HIDDEN_EXPORT_COLUMNS | {"Cliente"}),
        errors="ignore",
    )
    for column_name in DASHBOARD_RANKING_VISIBLE_COLUMNS:
        if column_name not in clean.columns:
            clean[column_name] = "" if column_name == "Client Name" else 0.0

    flag_columns = [
        col for col in ("__is_total__", "__is_grand_total__", "__is_highlight__")
        if col in clean.columns
    ]
    return clean[DASHBOARD_RANKING_VISIBLE_COLUMNS + flag_columns].copy()


def _dashboard_prepare_segment_dataframe(df: pd.DataFrame | None) -> pd.DataFrame:
    """Une Segmento y Región en una sola columna para la hoja Dashboard."""
    visible_columns = [
        "Segment / Region",
        "Actual",
        "Plan",
        "Fcst",
        "PY",
        "Var VS Plan",
        "%Var VS Plan",
        "Var VS Fcst",
        "%Var VS Fcst",
        "Var VS PY",
        "%Var VS PY",
    ]
    if df is None or df.empty:
        return pd.DataFrame(columns=visible_columns)

    clean = df.copy()
    labels = []
    for _, row in clean.iterrows():
        segment = str(row.get("Segmento", "")).strip()
        region = str(row.get("Región", "")).strip()
        is_grand_total = bool(row.get("__is_grand_total__", False))
        is_total = bool(row.get("__is_total__", False))

        if is_grand_total:
            label = "Total Mexico"
        elif is_total:
            label = f"{segment} | Total" if segment else "Total"
        elif segment and region:
            label = f"{segment} | {region}"
        else:
            label = segment or region
        labels.append(label)

    clean["Segment / Region"] = labels
    for column_name in visible_columns:
        if column_name not in clean.columns:
            clean[column_name] = None

    flag_columns = [
        col for col in ("__is_total__", "__is_grand_total__", "__is_highlight__")
        if col in clean.columns
    ]
    return clean[visible_columns + flag_columns].copy()


def _dashboard_write_title(
    worksheet,
    title: str,
    row: int,
    start_col: int,
    end_col: int,
) -> None:
    worksheet.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col,
    )
    cell = worksheet.cell(row=row, column=start_col, value=title)
    cell.font = DASHBOARD_TITLE_FONT
    cell.alignment = CENTER_ALIGNMENT
    cell.fill = PatternFill(fill_type="solid", fgColor=DASHBOARD_WHITE)
    cell.border = DASHBOARD_NO_BORDER
    worksheet.row_dimensions[row].height = 28


def _dashboard_write_period_header(
    worksheet,
    dashboard_tables: dict,
    start_row: int,
) -> int:
    month_label = dashboard_tables.get("month_label", "")
    latest_year = dashboard_tables.get("latest_year", "")
    currency_label = dashboard_tables.get("currency_label", "")

    period_rows = [
        ("Month", month_label),
        ("Year", latest_year),
    ]

    for offset, (label, value) in enumerate(period_rows):
        row = start_row + offset
        label_cell = worksheet.cell(row=row, column=1, value=label)
        value_cell = worksheet.cell(row=row, column=2, value=value)

        label_cell.font = DASHBOARD_PERIOD_LABEL_FONT
        value_cell.font = DASHBOARD_PERIOD_VALUE_FONT
        label_cell.alignment = LEFT_ALIGNMENT
        value_cell.alignment = LEFT_ALIGNMENT
        label_cell.border = DASHBOARD_NO_BORDER
        value_cell.border = DASHBOARD_NO_BORDER

    return start_row + 4


def _dashboard_write_section_title(
    worksheet,
    title: str,
    row: int,
    start_col: int,
    table_width: int,
    compact: bool,
) -> None:
    if compact:
        merge_width = min(max(3, table_width // 3), table_width)
        end_col = start_col + merge_width - 1
        fill = PatternFill(fill_type="solid", fgColor=DASHBOARD_TITLE_GRAY)
        border = Border(
            left=Side(style="thin", color=DASHBOARD_BLACK),
            right=Side(style="thin", color=DASHBOARD_BLACK),
            top=Side(style="thin", color=DASHBOARD_BLACK),
            bottom=Side(style="thin", color=DASHBOARD_BLACK),
        )
        alignment = CENTER_ALIGNMENT
    else:
        end_col = start_col + table_width - 1
        fill = PatternFill(fill_type="solid", fgColor=DASHBOARD_WHITE)
        border = DASHBOARD_NO_BORDER
        alignment = CENTER_ALIGNMENT

    worksheet.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col,
    )
    cell = worksheet.cell(row=row, column=start_col, value=title)
    cell.font = DASHBOARD_SECTION_FONT
    cell.alignment = alignment
    cell.fill = fill
    cell.border = border
    worksheet.row_dimensions[row].height = 18


DASHBOARD_FINAL_TOTAL_LABELS = {
    "total",
    "total general",
    "grand total",
    "total mexico",
    "total méxico",
}


def _dashboard_is_final_total_label(value) -> bool:
    return str(value or "").strip().lower() in DASHBOARD_FINAL_TOTAL_LABELS


def _dashboard_normalize_final_total_value(value):
    return "Total Mexico" if _dashboard_is_final_total_label(value) else value


def _dashboard_write_table(
    worksheet,
    original_df: pd.DataFrame | None,
    title: str,
    start_row: int,
    start_col: int,
    compact_title: bool = True,
    is_kpi: bool = False,
) -> tuple[int, int]:
    """
    Escribe una tabla con el diseño propio del Dashboard:
    fondo blanco, encabezados negros, líneas dobles y sin bordes verticales.
    """
    export_df = sanitize_export_dataframe(original_df)

    if export_df is None or export_df.empty:
        _dashboard_write_section_title(
            worksheet,
            title=title,
            row=start_row,
            start_col=start_col,
            table_width=3,
            compact=compact_title,
        )
        info_cell = worksheet.cell(
            row=start_row + 2,
            column=start_col,
            value="Sin información disponible",
        )
        info_cell.font = DASHBOARD_BODY_FONT
        info_cell.alignment = LEFT_ALIGNMENT
        info_cell.border = DASHBOARD_NO_BORDER
        return start_row + 2, start_col + 2

    num_cols = len(export_df.columns)
    title_row = start_row
    header_row = start_row + 1
    body_start_row = start_row + 2

    _dashboard_write_section_title(
        worksheet,
        title=title,
        row=title_row,
        start_col=start_col,
        table_width=num_cols,
        compact=compact_title,
    )

    for offset, column_name in enumerate(export_df.columns):
        current_col = start_col + offset
        cell = worksheet.cell(
            row=header_row,
            column=current_col,
            value=str(column_name),
        )
        cell.font = DASHBOARD_HEADER_FONT
        cell.fill = PatternFill(fill_type="solid", fgColor=DASHBOARD_WHITE)
        cell.alignment = CENTER_ALIGNMENT if offset == 0 else RIGHT_ALIGNMENT
        cell.border = (
            DASHBOARD_KPI_HEADER_BORDER if is_kpi else DASHBOARD_HEADER_BORDER
        )

    for row_index in range(len(export_df)):
        excel_row = body_start_row + row_index
        original_row = (
            original_df.iloc[row_index]
            if original_df is not None and row_index < len(original_df)
            else None
        )

        is_total = bool(
            original_row is not None
            and original_row.get("__is_total__", False)
        )
        is_grand_total = bool(
            original_row is not None
            and original_row.get("__is_grand_total__", False)
        )
        is_highlight = bool(
            original_row is not None
            and original_row.get("__is_highlight__", False)
        )

        first_label = str(export_df.iloc[row_index, 0]).strip()
        normalized_first_label = first_label.lower()

        # Regla única para todos los bloques del Dashboard:
        # el total final siempre se llama Total Mexico y siempre usa franja verde.
        # Los subtotales como "ACCO | Total" no entran en esta regla.
        if is_grand_total or _dashboard_is_final_total_label(normalized_first_label):
            is_grand_total = True
            is_total = False
            is_highlight = True

        kpi_label = normalized_first_label
        is_gsnr_row = is_kpi and kpi_label == "gsnr"
        is_achievement_row = is_kpi and "achievement" in kpi_label
        is_bts_row = is_kpi and kpi_label.startswith("bts")

        for col_index, column_name in enumerate(export_df.columns):
            excel_col = start_col + col_index
            raw_value = export_df.iloc[row_index, col_index]

            # La primera celda del total final se estandariza visualmente.
            if col_index == 0 and is_grand_total:
                raw_value = "Total Mexico"

            output_value = raw_value
            achievement_percent_cell = (
                is_achievement_row
                and str(column_name).strip() in {"Var VS Plan", "Var VS PY"}
            )
            if achievement_percent_cell:
                # El valor se conserva como razón (ej. 0.66) y Excel lo
                # presenta como porcentaje (66.00%), sin dividir entre 1,000.
                output_value = safe_float(raw_value)
            elif _dashboard_is_numeric_column(column_name):
                output_value = _dashboard_scale_value(column_name, raw_value)

            cell = worksheet.cell(
                row=excel_row,
                column=excel_col,
                value=output_value,
            )
            cell.font = (
                DASHBOARD_BODY_BOLD_FONT
                if col_index == 0 or is_total or is_grand_total or is_highlight or is_kpi
                else DASHBOARD_BODY_FONT
            )
            cell.alignment = (
                CENTER_ALIGNMENT
                if col_index == 0 and is_kpi
                else LEFT_ALIGNMENT
                if col_index == 0
                else RIGHT_ALIGNMENT
            )
            cell.border = DASHBOARD_NO_BORDER
            cell.fill = PatternFill(fill_type="solid", fgColor=DASHBOARD_WHITE)

            if is_gsnr_row:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=DASHBOARD_GSNR_GRAY,
                )
            elif is_bts_row:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=DASHBOARD_BTS_GRAY,
                )
            elif is_achievement_row:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=DASHBOARD_WHITE,
                )
            elif is_grand_total or is_highlight:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=DASHBOARD_GRAND_TOTAL_GREEN,
                )
                cell.border = DASHBOARD_TOTAL_BORDER
            elif is_total:
                cell.fill = PatternFill(
                    fill_type="solid",
                    fgColor=DASHBOARD_TOTAL_GRAY,
                )
                cell.border = DASHBOARD_TOTAL_BORDER

            if achievement_percent_cell:
                cell.number_format = DASHBOARD_PERCENT_FORMAT
                if is_negative_number(raw_value):
                    cell.font = DASHBOARD_NEGATIVE_FONT
            elif _dashboard_is_numeric_column(column_name):
                _dashboard_set_numeric_style(cell, column_name, raw_value)

        worksheet.row_dimensions[excel_row].height = 17

    # Anchos propios del Dashboard.
    for offset, column_name in enumerate(export_df.columns):
        excel_col = start_col + offset
        letter = get_column_letter(excel_col)

        if offset == 0:
            width = 28 if num_cols <= 9 else 20
        elif str(column_name).strip() in {"Cliente", "Client"}:
            width = 12
        elif _dashboard_is_percent_column(column_name):
            width = 13
        elif _dashboard_is_numeric_column(column_name):
            width = 12
        else:
            width = min(
                max(12, estimate_column_width(export_df[column_name], str(column_name))),
                26,
            )

        current_width = worksheet.column_dimensions[letter].width or 0
        worksheet.column_dimensions[letter].width = max(current_width, width)

    last_row = body_start_row + len(export_df) - 1
    last_col = start_col + num_cols - 1
    return last_row, last_col


def _dashboard_write_pair(
    worksheet,
    left_df: pd.DataFrame | None,
    right_df: pd.DataFrame | None,
    left_title: str,
    right_title: str,
    start_row: int,
    compact_title: bool = True,
    is_kpi: bool = False,
) -> tuple[int, int]:
    left_visible = sanitize_export_dataframe(left_df)
    left_width = max(len(left_visible.columns), 1)
    right_visible = sanitize_export_dataframe(right_df)
    right_width = max(len(right_visible.columns), 1)

    # Dos columnas de separación visual entre Monthly y YTD.
    right_start_col = 1 + left_width + 2

    left_last_row, left_last_col = _dashboard_write_table(
        worksheet=worksheet,
        original_df=left_df,
        title=left_title,
        start_row=start_row,
        start_col=1,
        compact_title=compact_title,
        is_kpi=is_kpi,
    )
    right_last_row, right_last_col = _dashboard_write_table(
        worksheet=worksheet,
        original_df=right_df,
        title=right_title,
        start_row=start_row,
        start_col=right_start_col,
        compact_title=compact_title,
        is_kpi=is_kpi,
    )

    return max(left_last_row, right_last_row), max(left_last_col, right_last_col)


def write_dashboard_tables_to_workbook(
    worksheet,
    dashboard_tables: dict,
) -> None:
    """
    Escribe el Dashboard tanto para la descarga individual como para la global.

    La estructura reproduce la vista de Streamlit:
    - Periodo en la parte superior izquierda.
    - Título rojo centrado.
    - KPIs Month/YTD lado a lado.
    - Cada reporte Monthly/YTD lado a lado.
    """
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A16"
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = 1
    worksheet.page_setup.fitToHeight = 0
    worksheet.sheet_properties.outlinePr.summaryBelow = True

    month_label = str(dashboard_tables.get("month_label", ""))
    report_title = str(
        dashboard_tables.get("dashboard_title")
        or getattr(config, "DASHBOARD_TITLE", "Mexico Dashboard 2026")
    )

    month_kpis = _dashboard_kpi_dataframe(
        dashboard_tables.get("client_table"),
        dashboard_tables.get("bts_table"),
        "MTD",
        f"BTS ({month_label})",
    )
    ytd_kpis = _dashboard_kpi_dataframe(
        dashboard_tables.get("client_table"),
        dashboard_tables.get("bts_table"),
        "YTD",
        f"BTS (Oct-{month_label})",
    )

    category_mtd, category_ytd = _dashboard_prepare_category_pair(
        dashboard_tables.get("category_mtd"),
        dashboard_tables.get("category_ytd"),
    )
    ranking_mtd = _dashboard_prepare_ranking_dataframe(
        dashboard_tables.get("ranking_mtd")
    )
    ranking_ytd = _dashboard_prepare_ranking_dataframe(
        dashboard_tables.get("ranking_ytd")
    )
    segment_mtd = _dashboard_prepare_segment_dataframe(
        dashboard_tables.get("segment_mtd")
    )
    segment_ytd = _dashboard_prepare_segment_dataframe(
        dashboard_tables.get("segment_ytd")
    )

    # Calcula el ancho máximo necesario antes de centrar el título.
    all_pairs = [
        (month_kpis, ytd_kpis),
        (dashboard_tables.get("report1_mtd"), dashboard_tables.get("report1_ytd")),
        (segment_mtd, segment_ytd),
        (category_mtd, category_ytd),
        (dashboard_tables.get("channel_mtd"), dashboard_tables.get("channel_ytd")),
        (ranking_mtd, ranking_ytd),
    ]
    max_end_col = 18
    for left_df, right_df in all_pairs:
        left_width = max(len(sanitize_export_dataframe(left_df).columns), 1)
        right_width = max(len(sanitize_export_dataframe(right_df).columns), 1)
        max_end_col = max(max_end_col, left_width + 2 + right_width)

    current_row = _dashboard_write_period_header(
        worksheet,
        dashboard_tables,
        start_row=2,
    )

    _dashboard_write_title(
        worksheet=worksheet,
        title=report_title,
        row=current_row,
        start_col=1,
        end_col=max_end_col,
    )
    current_row += 3

    currency_cell = worksheet.cell(
        row=current_row,
        column=2,
        value=dashboard_tables.get("currency_label", ""),
    )
    currency_cell.font = DASHBOARD_BODY_BOLD_FONT
    currency_cell.alignment = LEFT_ALIGNMENT
    current_row += 2

    last_row, _ = _dashboard_write_pair(
        worksheet=worksheet,
        left_df=month_kpis,
        right_df=ytd_kpis,
        left_title="Sales Month",
        right_title="Sales YTD",
        start_row=current_row,
        compact_title=False,
        is_kpi=True,
    )
    current_row = last_row + 3

    dashboard_sections = [
        (
            "Sales by Channel Monthly",
            "Sales by Channel YTD",
            dashboard_tables.get("report1_mtd"),
            dashboard_tables.get("report1_ytd"),
        ),
        (
            "Segment x Region Monthly",
            "Segment x Region YTD",
            segment_mtd,
            segment_ytd,
        ),
        (
            "Sales by Category Monthly",
            "Sales by Category YTD",
            category_mtd,
            category_ytd,
        ),
        (
            "Channel Monthly",
            "Channel YTD",
            dashboard_tables.get("channel_mtd"),
            dashboard_tables.get("channel_ytd"),
        ),
        (
            "Ranking Clientes Monthly",
            "Ranking Clientes YTD",
            ranking_mtd,
            ranking_ytd,
        ),
    ]

    for left_title, right_title, left_df, right_df in dashboard_sections:
        last_row, _ = _dashboard_write_pair(
            worksheet=worksheet,
            left_df=left_df,
            right_df=right_df,
            left_title=left_title,
            right_title=right_title,
            start_row=current_row,
            compact_title=True,
            is_kpi=False,
        )
        current_row = last_row + 3

    # Vista inicial cómoda al abrir el archivo.
    worksheet.sheet_view.zoomScale = 80
    worksheet.sheet_view.zoomScaleNormal = 80
    worksheet.print_options.horizontalCentered = True
    worksheet.sheet_view.showRowColHeaders = True


def build_dashboard_excel_bytes(
    dashboard_tables: dict,
    report_title: str | None = None,
    sheet_name: str | None = None,
    progress_callback=None,
) -> bytes:
    total_steps = 4
    emit_progress(
        progress_callback,
        "Preparando la exportación del Dashboard",
        1,
        total_steps,
    )
    output, writer = create_excel_writer_buffer()

    emit_progress(
        progress_callback,
        "Creando la hoja del Dashboard",
        2,
        total_steps,
    )
    worksheet = ensure_sheet_exists(
        writer,
        sheet_name or getattr(config, "EXPORT_SHEET_DASHBOARD", "Dashboard"),
    )

    tables = dict(dashboard_tables or {})
    # El título del archivo debe conservar exactamente el título visual
    # del Dashboard, no el título contextual usado por otros reportes.
    tables["dashboard_title"] = getattr(
        config,
        "DASHBOARD_TITLE",
        "Mexico Dashboard 2026",
    )

    emit_progress(
        progress_callback,
        "Reproduciendo el diseño ejecutivo del Dashboard",
        3,
        total_steps,
    )
    write_dashboard_tables_to_workbook(worksheet, tables)

    emit_progress(
        progress_callback,
        "Finalizando el archivo Excel del Dashboard",
        4,
        total_steps,
    )
    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)


# ---------------------------------------------------------
# EXPORTACIÓN GLOBAL
# ---------------------------------------------------------
def build_full_reports_excel_bytes(
    base_mtd_tables: dict | None = None,
    report_1_tables: dict | None = None,
    report_2_segment_tables: dict | None = None,
    report_2_category_tables: dict | None = None,
    report_3_tables: dict | None = None,
    report_4_tables: dict | None = None,
    dashboard_tables: dict | None = None,
    progress_callback=None,
) -> bytes:
    selected_sections = [
        ("Base MTD", base_mtd_tables),
        ("Reporte 1 - Oficina de ventas", report_1_tables),
        ("Reporte 2 - Segment x Region", report_2_segment_tables),
        ("Reporte 2 - Category", report_2_category_tables),
        ("Reporte 3 - Channel", report_3_tables),
        ("Reporte 4 - Ranking de Clientes", report_4_tables),
        ("Dashboard", dashboard_tables),
    ]
    active_sections = [name for name, tables in selected_sections if tables]
    total_steps = len(active_sections) + 2

    emit_progress(progress_callback, "Preparando la descarga global de reportes", 1, total_steps)

    output, writer = create_excel_writer_buffer()
    workbook_created = False
    current_step = 2

    if base_mtd_tables:
        emit_progress(progress_callback, "Agregando Base MTD al archivo global", current_step, total_steps)
        current_step += 1
        ws = ensure_sheet_exists(
            writer,
            getattr(config, "EXPORT_SHEET_BASE_MTD", "Base MTD"),
        )

        write_base_mtd_tables_to_workbook(
            worksheet=ws,
            base_mtd_tables=base_mtd_tables,
        )

        workbook_created = True

    if report_1_tables:
        emit_progress(progress_callback, "Agregando Reporte 1 - Oficina de ventas", current_step, total_steps)
        current_step += 1
        ws = ensure_sheet_exists(writer, "Reporte 1")

        current_row = write_global_title(
            worksheet=ws,
            report_title=get_report_title_from_tables(report_1_tables),
            start_row=1,
            start_col=1,
            width=8,
        )

        write_two_tables_side_by_side(
            worksheet=ws,
            left_df=report_1_tables["mtd_without_kens"],
            right_df=report_1_tables["ytd_without_kens"],
            left_title="MTD Oficina de ventas",
            right_title="YTD Oficina de ventas",
            start_row=current_row,
        )

        workbook_created = True

    if report_2_segment_tables:
        emit_progress(progress_callback, "Agregando Reporte 2 - Segment x Region", current_step, total_steps)
        current_step += 1
        ws = ensure_sheet_exists(writer, "Reporte 2 - Segment")

        current_row = write_global_title(
            worksheet=ws,
            report_title=get_report_title_from_tables(report_2_segment_tables),
            start_row=1,
            start_col=1,
            width=8,
        )

        write_two_tables_side_by_side(
            worksheet=ws,
            left_df=report_2_segment_tables["mtd"],
            right_df=report_2_segment_tables["ytd"],
            left_title="MTD Segment x Region",
            right_title="YTD Segment x Region",
            start_row=current_row,
        )

        workbook_created = True

    if report_2_category_tables:
        emit_progress(progress_callback, "Agregando Reporte 2 - Category", current_step, total_steps)
        current_step += 1
        ws = ensure_sheet_exists(writer, "Reporte 2 - Category")

        current_row = write_global_title(
            worksheet=ws,
            report_title=get_report_title_from_tables(report_2_category_tables),
            start_row=1,
            start_col=1,
            width=9,
        )

        write_two_tables_side_by_side(
            worksheet=ws,
            left_df=report_2_category_tables["mtd"],
            right_df=report_2_category_tables["ytd"],
            left_title="MTD Category",
            right_title="YTD Category",
            start_row=current_row,
        )

        workbook_created = True

    if report_3_tables:
        emit_progress(progress_callback, "Agregando Reporte 3 - Channel", current_step, total_steps)
        current_step += 1
        ws = ensure_sheet_exists(writer, "Reporte 3")

        current_row = write_global_title(
            worksheet=ws,
            report_title=get_report_title_from_tables(report_3_tables),
            start_row=1,
            start_col=1,
            width=8,
        )

        write_two_tables_side_by_side(
            worksheet=ws,
            left_df=report_3_tables["mtd"],
            right_df=report_3_tables["ytd"],
            left_title="MTD Channel",
            right_title="YTD Channel",
            start_row=current_row,
        )

        workbook_created = True

    if report_4_tables:
        emit_progress(progress_callback, "Agregando Reporte 4 - Ranking de Clientes", current_step, total_steps)
        current_step += 1
        ws = ensure_sheet_exists(writer, "Reporte 4")

        current_row = write_global_title(
            worksheet=ws,
            report_title=get_report_title_from_tables(report_4_tables),
            start_row=1,
            start_col=1,
            width=len(REPORT_4_VISIBLE_COLUMNS),
        )

        write_two_tables_side_by_side(
            worksheet=ws,
            left_df=prepare_report_4_export_dataframe(report_4_tables["mtd"]),
            right_df=prepare_report_4_export_dataframe(report_4_tables["ytd"]),
            left_title="MTD Ranking Clients",
            right_title="YTD Ranking Clients",
            start_row=current_row,
        )

        workbook_created = True

    if dashboard_tables:
        emit_progress(progress_callback, "Agregando Dashboard ejecutivo", current_step, total_steps)
        current_step += 1
        ws = ensure_sheet_exists(
            writer,
            getattr(config, "EXPORT_SHEET_DASHBOARD", "Dashboard"),
        )
        write_dashboard_tables_to_workbook(ws, dashboard_tables)
        workbook_created = True

    if not workbook_created:
        ws = ensure_sheet_exists(writer, "Reportes")
        ws.cell(row=1, column=1, value="No hay reportes construidos para exportar.")

    remove_default_sheet_if_needed(writer)
    return build_excel_bytes_from_writer(output, writer)

# =========================================================
# INTEGRACIÓN FORECAST - EXTENSIÓN DE EXPORTACIONES
# =========================================================
HEADER_FCST_FILL = PatternFill(fill_type="solid", fgColor="34A853")
HEADER_FCST_SKU_FILL = PatternFill(fill_type="solid", fgColor="FFC34D")
HEADER_VAR_PLAN_FILL = PatternFill(fill_type="solid", fgColor="F4B400")
HEADER_VAR_FCST_FILL = PatternFill(fill_type="solid", fgColor="34A853")
HEADER_VAR_PY_FILL = PatternFill(fill_type="solid", fgColor="0B5A7A")

PERCENT_COLUMNS.add("%Var VS Fcst")
NUMERIC_COLUMNS.update({"Fcst", "Var VS Fcst"})

# Ranking: conserva exactamente el orden visual acordado.
REPORT_4_VISIBLE_COLUMNS = [
    "TOP",
    "Client Name",
    "Cliente",
    "Actual",
    "Plan",
    "Var VS Plan",
    "%Var VS Plan",
    "Fcst",
    "Var VS Fcst",
    "%Var VS Fcst",
    "PY",
    "Var VS PY",
    "%Var VS PY",
]

def get_header_fill(column_name: str):
    name = str(column_name).strip()

    if name == "Actual":
        return HEADER_ACTUAL_FILL
    if name == "Plan":
        return HEADER_PLAN_FILL
    if name == "Fcst":
        return HEADER_FCST_FILL
    if name == "PY":
        return HEADER_PY_FILL
    if name in {"Var VS Plan", "%Var VS Plan"}:
        return HEADER_VAR_PLAN_FILL
    if name in {"Var VS Fcst", "%Var VS Fcst"}:
        return HEADER_VAR_FCST_FILL
    if name in {"Var VS PY", "%Var VS PY"}:
        return HEADER_VAR_PY_FILL

    return HEADER_NEUTRAL_FILL

